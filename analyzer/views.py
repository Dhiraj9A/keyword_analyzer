# import re

# from collections import Counter

# from bs4 import BeautifulSoup

# from django.shortcuts import render

# from .services.fetcher import fetch_website

# from openpyxl import Workbook

# def home(request):

#     return render(
#         request,
#         "analyzer/home.html"
#     )


# def analyze_url(request):

#     if request.method != "POST":

#         return render(
#             request,
#             "analyzer/home.html"
#         )

#     url = request.POST.get(
#         "url",
#         ""
#     ).strip()

#     # ==========================================
#     # URL VALIDATION
#     # ==========================================

#     if not url:

#         return render(
#             request,
#             "analyzer/home.html",
#             {
#                 "error": "Please enter a website URL."
#             }
#         )

#     if not url.startswith(
#         ("http://", "https://")
#     ):

#         url = "https://" + url

#     # ==========================================
#     # FETCH WEBSITE
#     # ==========================================

#     result = fetch_website(url)

#     if not result["success"]:

#         return render(
#             request,
#             "analyzer/home.html",
#             {
#                 "error": (
#                     "Unable to analyze website."
#                 ),
#                 "details": result["error"],
#             }
#         )

#     html = result["html"]

#     fetch_method = result["method"]

#     # ==========================================
#     # PARSE HTML
#     # ==========================================

#     soup = BeautifulSoup(
#         html,
#         "html.parser"
#     )

#     # ==========================================
#     # REMOVE UNWANTED ELEMENTS
#     # ==========================================

#     for element in soup.find_all(
#         [
#             "script",
#             "style",
#             "noscript",
#             "svg",
#         ]
#     ):

#         element.decompose()

#     # ==========================================
#     # TITLE
#     # ==========================================

#     title = ""

#     if soup.title:

#         title = soup.title.get_text(
#             strip=True
#         )

#     # ==========================================
#     # META DESCRIPTION
#     # ==========================================

#     meta_description = ""

#     meta = soup.find(
#         "meta",
#         attrs={
#             "name": re.compile(
#                 "^description$",
#                 re.I
#             )
#         }
#     )

#     if meta:

#         meta_description = meta.get(
#             "content",
#             ""
#         ).strip()

#     # ==========================================
#     # HEADINGS
#     # ==========================================

#     headings = []

#     for tag in soup.find_all(
#         ["h1", "h2", "h3"]
#     ):

#         heading_text = tag.get_text(
#             " ",
#             strip=True
#         )

#         if heading_text:

#             headings.append(
#                 {
#                     "tag": tag.name.upper(),
#                     "text": heading_text,
#                 }
#             )

#     # ==========================================
#     # WEBSITE TEXT
#     # ==========================================

#     text = soup.get_text(
#         " ",
#         strip=True
#     )

#     text = text.lower()

#     # ==========================================
#     # WORD EXTRACTION
#     # ==========================================

#     words = re.findall(
#         r"\b[a-zA-Z]{3,}\b",
#         text
#     )

#     # ==========================================
#     # STOPWORDS
#     # ==========================================

#     stopwords = {
#         "the",
#         "and",
#         "for",
#         "with",
#         "that",
#         "this",
#         "from",
#         "your",
#         "you",
#         "are",
#         "was",
#         "were",
#         "have",
#         "has",
#         "had",
#         "will",
#         "can",
#         "our",
#         "about",
#         "into",
#         "more",
#         "than",
#         "their",
#         "they",
#         "them",
#         "then",
#         "also",
#         "not",
#         "but",
#         "all",
#         "www",
#         "http",
#         "https",
#         "com",
#         "html",
#         "home",
#         "page",
#         "click",
#         "here",
#     }

#     # ==========================================
#     # FILTER WORDS
#     # ==========================================

#     filtered_words = [
#         word
#         for word in words
#         if word not in stopwords
#     ]

#     # ==========================================
#     # KEYWORD COUNT
#     # ==========================================

#     keyword_counter = Counter(
#         filtered_words
#     )

#     keywords = keyword_counter.most_common(
#         50
#     )

#     # ==========================================
#     # RESULT
#     # ==========================================

#     context = {
#         "url": url,

#         "title": title,

#         "meta_description": (
#             meta_description
#         ),

#         "headings": headings,

#         "keywords": keywords,

#         "word_count": len(
#             filtered_words
#         ),

#         "fetch_method": fetch_method,
#     }

#     return render(
#         request,
#         "analyzer/results.html",
#         context
#     )


#     #________________ Exel ______________________#



import re

from collections import Counter

from bs4 import BeautifulSoup

from django.shortcuts import render
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


from .services.fetcher import fetch_website

# =====================================================
# HOME
# =====================================================

def home(request):

    return render(
        request,
        "analyzer/home.html"
    )


# =====================================================
# ANALYZE URL
# =====================================================

def analyze_url(request):

    if request.method != "POST":

        return render(
            request,
            "analyzer/home.html"
        )

    url = request.POST.get(
        "url",
        ""
    ).strip()

    # =================================================
    # URL VALIDATION
    # =================================================

    if not url:

        return render(
            request,
            "analyzer/home.html",
            {
                "error": "Please enter a website URL."
            }
        )

    if not url.startswith(
        ("http://", "https://")
    ):

        url = "https://" + url

    # =================================================
    # FETCH WEBSITE
    # =================================================

    result = fetch_website(url)

    if not result["success"]:

        return render(
            request,
            "analyzer/home.html",
            {
                "error": "Unable to analyze website.",
                "details": result["error"],
            }
        )

    html = result["html"]

    fetch_method = result["method"]

    # =================================================
    # PARSE HTML
    # =================================================

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    # =================================================
    # REMOVE UNWANTED ELEMENTS
    # =================================================

    for element in soup.find_all(
        [
            "script",
            "style",
            "noscript",
            "svg",
        ]
    ):

        element.decompose()

    # =================================================
    # TITLE
    # =================================================

    title = ""

    if soup.title:

        title = soup.title.get_text(
            strip=True
        )

    # =================================================
    # META DESCRIPTION
    # =================================================

    meta_description = ""

    meta = soup.find(
        "meta",
        attrs={
            "name": re.compile(
                "^description$",
                re.I
            )
        }
    )

    if meta:

        meta_description = meta.get(
            "content",
            ""
        ).strip()

    # =================================================
    # HEADINGS
    # =================================================

    headings = []

    for tag in soup.find_all(
        ["h1", "h2", "h3"]
    ):

        heading_text = tag.get_text(
            " ",
            strip=True
        )

        if heading_text:

            headings.append(
                {
                    "tag": tag.name.upper(),
                    "text": heading_text,
                }
            )

    # =================================================
    # WEBSITE TEXT
    # =================================================

    text = soup.get_text(
        " ",
        strip=True
    )

    text = text.lower()

    # =================================================
    # WORD EXTRACTION
    # =================================================

    words = re.findall(
        r"\b[a-zA-Z]{3,}\b",
        text
    )

    # =================================================
    # STOPWORDS
    # =================================================

    stopwords = {
        "the",
        "and",
        "for",
        "with",
        "that",
        "this",
        "from",
        "your",
        "you",
        "are",
        "was",
        "were",
        "have",
        "has",
        "had",
        "will",
        "can",
        "our",
        "about",
        "into",
        "more",
        "than",
        "their",
        "they",
        "them",
        "then",
        "also",
        "not",
        "but",
        "all",
        "www",
        "http",
        "https",
        "com",
        "html",
        "home",
        "page",
        "click",
        "here",
    }

    # =================================================
    # FILTER WORDS
    # =================================================

    filtered_words = [
        word
        for word in words
        if word not in stopwords
    ]

    # =================================================
    # KEYWORD COUNT
    # =================================================

    keyword_counter = Counter(
        filtered_words
    )

    keywords = keyword_counter.most_common(
        50
    )

    # =================================================
    # RESULT CONTEXT
    # =================================================

    context = {

        "url": url,

        "title": title,

        "meta_description": (
            meta_description
        ),

        "headings": headings,

        "keywords": keywords,

        "word_count": len(
            filtered_words
        ),

        "fetch_method": fetch_method,

    }

    # =================================================
    # SHOW RESULT
    # =================================================

    return render(
        request,
        "analyzer/results.html",
        context
    )

#  # =====================================================
# # DOWNLOAD EXCEL
# # =====================================================

def download_excel(request):

    if request.method != "POST":
        return render(
            request,
            "analyzer/home.html",
            {
                "error": "Invalid request."
            }
        )

    # =====================================================
    # URL
    # =====================================================

    url = request.POST.get(
        "url",
        ""
    ).strip()

    if not url:

        return render(
            request,
            "analyzer/home.html",
            {
                "error": "Website URL is required."
            }
        )

    if not url.startswith(
        ("http://", "https://")
    ):
        url = "https://" + url

    # =====================================================
    # FETCH WEBSITE
    # =====================================================

    result = fetch_website(url)

    if not result["success"]:

        return render(
            request,
            "analyzer/home.html",
            {
                "error": "Unable to analyze website.",
                "details": result["error"],
            }
        )

    html = result["html"]

    fetch_method = result["method"]

    # =====================================================
    # PARSE HTML
    # =====================================================

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    # =====================================================
    # REMOVE UNWANTED ELEMENTS
    # =====================================================

    for element in soup.find_all(
        [
            "script",
            "style",
            "noscript",
            "svg"
        ]
    ):
        element.decompose()

    # =====================================================
    # TITLE
    # =====================================================

    title = ""

    if soup.title:

        title = soup.title.get_text(
            strip=True
        )

    # =====================================================
    # META DESCRIPTION
    # =====================================================

    meta_description = ""

    meta = soup.find(
        "meta",
        attrs={
            "name": re.compile(
                "^description$",
                re.I
            )
        }
    )

    if meta:

        meta_description = meta.get(
            "content",
            ""
        ).strip()

    # =====================================================
    # HEADINGS
    # =====================================================

    headings = []

    for tag in soup.find_all(
        [
            "h1",
            "h2",
            "h3",
            "h4",
            "h5",
            "h6"
        ]
    ):

        heading_text = tag.get_text(
            " ",
            strip=True
        )

        if heading_text:

            headings.append(
                {
                    "tag": tag.name.upper(),
                    "text": heading_text
                }
            )

    # =====================================================
    # WEBSITE TEXT
    # =====================================================

    text = soup.get_text(
        " ",
        strip=True
    ).lower()

    # =====================================================
    # WORD EXTRACTION
    # =====================================================

    words = re.findall(
        r"\b[a-zA-Z]{3,}\b",
        text
    )

    # =====================================================
    # STOPWORDS
    # =====================================================

    stopwords = {
        "the",
        "and",
        "for",
        "with",
        "that",
        "this",
        "from",
        "your",
        "you",
        "are",
        "was",
        "were",
        "have",
        "has",
        "had",
        "will",
        "can",
        "our",
        "about",
        "into",
        "more",
        "than",
        "their",
        "they",
        "them",
        "then",
        "also",
        "not",
        "but",
        "all",
        "www",
        "http",
        "https",
        "com",
        "html",
        "home",
        "page",
        "click",
        "here"
    }

    # =====================================================
    # FILTER WORDS
    # =====================================================

    filtered_words = [
        word
        for word in words
        if word not in stopwords
    ]

    # =====================================================
    # KEYWORDS
    # =====================================================

    keyword_counter = Counter(
        filtered_words
    )

    keywords = keyword_counter.most_common(
        50
    )

    total_words = len(
        filtered_words
    )

    # =====================================================
    # CREATE WORKBOOK
    # =====================================================

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "SEO Analysis"

    # =====================================================
    # BORDER
    # =====================================================

    thin_side = Side(
        style="thin",
        color="000000"
    )

    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    # =====================================================
    # FONTS
    # =====================================================

    section_font = Font(
        bold=True,
        size=14
    )

    header_font = Font(
        bold=True,
        size=11
    )

    normal_font = Font(
        size=10
    )

    # =====================================================
    # ALIGNMENTS
    # =====================================================

    center_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    left_alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    # =====================================================
    # COLUMN WIDTH
    # =====================================================

    worksheet.column_dimensions[
        "A"
    ].width = 22

    worksheet.column_dimensions[
        "B"
    ].width = 35

    worksheet.column_dimensions[
        "C"
    ].width = 55

    worksheet.column_dimensions[
        "D"
    ].width = 20

    # =====================================================
    # SECTION 1
    # WEBSITE SEO ANALYSIS
    # =====================================================

    worksheet.merge_cells(
        "A1:D1"
    )

    worksheet["A1"] = (
        "WEBSITE SEO ANALYSIS"
    )

    worksheet["A1"].font = section_font

    worksheet["A1"].alignment = (
        center_alignment
    )

    worksheet["A1"].border = (
        thin_border
    )

    worksheet.row_dimensions[
        1
    ].height = 28

    # Apply border to merged cells

    for cell in worksheet["1:1"]:

        cell.border = thin_border

    # =====================================================
    # FIELD / VALUE
    # =====================================================

    worksheet["A3"] = "Field"
    worksheet["B3"] = "Value"

    worksheet["A3"].font = header_font
    worksheet["B3"].font = header_font

    worksheet["A3"].alignment = left_alignment
    worksheet["B3"].alignment = left_alignment

    worksheet["A3"].border = thin_border
    worksheet["B3"].border = thin_border

    # Keep C and D bordered

    worksheet["C3"].border = thin_border
    worksheet["D3"].border = thin_border

    # =====================================================
    # SEO INFORMATION
    # =====================================================

    seo_data = [

        (
            "Website URL",
            url
        ),

        (
            "Title",
            title
        ),

        (
            "Meta Description",
            meta_description
        ),

        (
            "Total Words",
            total_words
        ),

        (
            "Fetch Method",
            fetch_method
        ),

        (
            "Total Headings",
            len(headings)
        ),

        (
            "Total Keywords",
            len(keywords)
        )

    ]

    row = 4

    for field, value in seo_data:

        worksheet.cell(
            row=row,
            column=1,
            value=field
        )

        worksheet.cell(
            row=row,
            column=2,
            value=value
        )

        # Border A:D

        for column in range(
            1,
            5
        ):

            cell = worksheet.cell(
                row=row,
                column=column
            )

            cell.border = thin_border

            cell.alignment = left_alignment

            cell.font = normal_font

        row += 1

    # =====================================================
    # WEBSITE HEADINGS TITLE
    # =====================================================

    headings_title_row = (
        row + 1
    )

    worksheet.merge_cells(
        start_row=headings_title_row,
        start_column=1,
        end_row=headings_title_row,
        end_column=4
    )

    worksheet.cell(
        row=headings_title_row,
        column=1,
        value="WEBSITE HEADINGS"
    )

    worksheet.cell(
        row=headings_title_row,
        column=1
    ).font = section_font

    worksheet.cell(
        row=headings_title_row,
        column=1
    ).alignment = center_alignment

    worksheet.row_dimensions[
        headings_title_row
    ].height = 28

    # =====================================================
    # BORDER FOR HEADING SECTION TITLE
    # =====================================================

    for column in range(
        1,
        5
    ):

        cell = worksheet.cell(
            row=headings_title_row,
            column=column
        )

        cell.border = thin_border

    # =====================================================
    # HEADINGS HEADER
    # =====================================================

    headings_header_row = (
        headings_title_row + 2
    )

    headings_headers = [
        "#",
        "Tag",
        "Heading",
        ""
    ]

    for column, value in enumerate(
        headings_headers,
        start=1
    ):

        cell = worksheet.cell(
            row=headings_header_row,
            column=column,
            value=value
        )

        cell.font = header_font

        cell.alignment = left_alignment

        cell.border = thin_border

    # =====================================================
    # HEADINGS DATA
    # =====================================================

    heading_row = (
        headings_header_row + 1
    )

    for index, heading in enumerate(
        headings,
        start=1
    ):

        values = [
            index,
            heading["tag"],
            heading["text"],
            ""
        ]

        for column, value in enumerate(
            values,
            start=1
        ):

            cell = worksheet.cell(
                row=heading_row,
                column=column,
                value=value
            )

            cell.border = thin_border

            cell.alignment = left_alignment

            cell.font = normal_font

        heading_row += 1

    # =====================================================
    # TOP KEYWORDS TITLE
    # =====================================================

    keywords_title_row = (
        heading_row + 1
    )

    worksheet.merge_cells(
        start_row=keywords_title_row,
        start_column=1,
        end_row=keywords_title_row,
        end_column=4
    )

    worksheet.cell(
        row=keywords_title_row,
        column=1,
        value="TOP KEYWORDS"
    )

    worksheet.cell(
        row=keywords_title_row,
        column=1
    ).font = section_font

    worksheet.cell(
        row=keywords_title_row,
        column=1
    ).alignment = center_alignment

    worksheet.row_dimensions[
        keywords_title_row
    ].height = 28

    # =====================================================
    # BORDER FOR KEYWORD TITLE
    # =====================================================

    for column in range(
        1,
        5
    ):

        cell = worksheet.cell(
            row=keywords_title_row,
            column=column
        )

        cell.border = thin_border

    # =====================================================
    # KEYWORD HEADER
    # =====================================================

    keyword_header_row = (
        keywords_title_row + 2
    )

    keyword_headers = [
        "#",
        "Keyword",
        "Frequency",
        "Density (%)"
    ]

    for column, value in enumerate(
        keyword_headers,
        start=1
    ):

        cell = worksheet.cell(
            row=keyword_header_row,
            column=column,
            value=value
        )

        cell.font = header_font

        cell.alignment = left_alignment

        cell.border = thin_border

    # =====================================================
    # KEYWORD DATA
    # =====================================================

    keyword_row = (
        keyword_header_row + 1
    )

    for index, (
        keyword,
        count
    ) in enumerate(
        keywords,
        start=1
    ):

        if total_words > 0:

            density = (
                count /
                total_words
            ) * 100

        else:

            density = 0

        values = [
            index,
            keyword,
            count,
            round(
                density,
                2
            )
        ]

        for column, value in enumerate(
            values,
            start=1
        ):

            cell = worksheet.cell(
                row=keyword_row,
                column=column,
                value=value
            )

            cell.border = thin_border

            cell.alignment = left_alignment

            cell.font = normal_font

        keyword_row += 1

    # =====================================================
    # FREEZE
    # =====================================================

    worksheet.freeze_panes = "A3"

    # =====================================================
    # PAGE SETUP
    # =====================================================

    worksheet.sheet_view.showGridLines = False

    # =====================================================
    # DOWNLOAD
    # =====================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="website_seo_analysis.xlsx"'
    )

    workbook.save(
        response
    )

    return response